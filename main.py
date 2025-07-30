import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import PatternFill
from openpyxl.utils import get_column_letter
import tkinter as tk
from tkinter import filedialog, messagebox
import os

def select_file(title, file_description):
    """Helper function to select a file using dialog"""
    print(f"Please select your Excel file ({file_description})...")
    file_path = filedialog.askopenfilename(
        title=title,
        filetypes=[
            ("Excel files", "*.xlsx *.xls"),
            ("All files", "*.*")
        ],
        initialdir=os.getcwd()
    )
    
    if not file_path:
        print("No file selected. Exiting...")
        return None
    
    print(f"Selected file: {file_path}")
    return file_path

def main():
    print("=== Attendance and Overtime Cell Coloring Tool ===\n")
    
    # Initialize tkinter root (hidden window)
    root = tk.Tk()
    root.withdraw()  # Hide the main tkinter window
    
    try:
        # First Excel File - Attendance data
        file_path = select_file("Select Attendance Excel File", "Attendance")
        if not file_path:
            return None
            
        df = pd.read_excel(file_path)
        print(f"Successfully loaded attendance file with {len(df)} rows and {len(df.columns)} columns.")
        
        absentees_by_date = {}
        
        # Date columns (from column 11 onwards)
        attendance_columns = df.columns[11:]
        
        print("Processing attendance data...")
        for date in attendance_columns:
            d = pd.to_datetime(str(date), errors='coerce', dayfirst=False)
            if pd.isna(d):
                formatted_date = str(date)
            else:
                formatted_date = d.strftime("%d/%m/%Y") 
                
            absent_mask = df[date].isna() | (df[date].astype(str).str.upper().isin(["A", "HN"]))
            absent_rows = df[absent_mask]
            
            absent_list = absent_rows['Employee ID'].tolist()
            
            absentees_by_date[formatted_date] = absent_list
        
        print(f"Found {len(absentees_by_date)} dates with attendance data")
        
        # Second Excel File - OT file
        overtime_file_path = select_file("Select Overtime Excel File", "Overtime")
        if not overtime_file_path:
            return None
            
        df_overtime = pd.read_excel(overtime_file_path)
        print(f"Successfully loaded overtime file with {len(df_overtime)} rows and {len(df_overtime.columns)} columns.")
        
        # Create row dictionary (Employee ID -> Row number)
        row_dictionary = {str(int(emp_id)): idx + 2 for idx, emp_id in enumerate(df_overtime['Emp_Id'])}
        
        # Create column dictionary (Date -> Column letter)
        column_dictionary = {}
        for idx, col in enumerate(df_overtime.columns):
            if col == 'Emp_Id': 
                continue
            
            d = pd.to_datetime(str(col), errors='coerce', dayfirst=False)
            
            if not pd.isna(d):
                formatted_date = d.strftime("%d/%m/%Y")
                column_letter = get_column_letter(idx + 1)
                column_dictionary[formatted_date] = column_letter
        
        # Load the workbook and worksheet
        print("Processing overtime data and applying colors...")
        workbook = load_workbook(overtime_file_path)
        worksheet = workbook.active
        
        # Define fill colors
        red_fill = PatternFill(start_color='FF0000', end_color='FF0000', fill_type='solid')  # Red
        green_fill = PatternFill(start_color='00FF00', end_color='00FF00', fill_type='solid')  # Green
        
        # Counter for tracking processed cells
        processed_cells = 0
        red_cells = 0
        green_cells = 0
        
        # Convert master dictionary to match the expected format (date -> list of employee IDs)
        master_dict = {}
        for date, emp_list in absentees_by_date.items():
            for emp_id in emp_list:
                if date not in master_dict:
                    master_dict[date] = []
                master_dict[date].append(str(emp_id))
        
        # Iterate through master dictionary (absent employees)
        for date, employee_ids in master_dict.items():
            for employee_id in employee_ids:
                # Check if the absent employee is present in the row dictionary
                if employee_id in row_dictionary:
                    # Check if the date is present in the column dictionary
                    if date in column_dictionary:
                        # Get row number and column letter
                        row_number = row_dictionary[employee_id]
                        column_letter = column_dictionary[date]
                        
                        # Get the cell value
                        cell = worksheet[f"{column_letter}{row_number}"]
                        cell_value = cell.value
                        
                        # Convert cell value to number if it's not None
                        try:
                            if cell_value is None:
                                cell_value = 0
                            else:
                                cell_value = float(cell_value)
                        except (ValueError, TypeError):
                            cell_value = 0
                        
                        # Apply color on the value
                        if cell_value > 0:
                            cell.fill = red_fill
                            red_cells += 1
                        else:
                            cell.fill = green_fill
                            green_cells += 1
                        
                        processed_cells += 1
        
        # Generate output filename
        if overtime_file_path.endswith('.xlsx'):
            output_path = overtime_file_path.replace('.xlsx', '_colored.xlsx')
        elif overtime_file_path.endswith('.xls'):
            output_path = overtime_file_path.replace('.xls', '_colored.xlsx')
        else:
            output_path = overtime_file_path + '_colored.xlsx'
        
        # Save the modified workbook
        print("Saving the modified file...")
        workbook.save(output_path)
        workbook.close()
        
        print(f"\n=== Processing Complete! ===")
        print(f"Total cells processed: {processed_cells}")
        print(f"Red cells (Absent + Overtime > 0): {red_cells}")
        print(f"Green cells (Absent + Overtime = 0): {green_cells}")
        print(f"Modified file saved as: {output_path}")
        
        # Show success message box
        messagebox.showinfo(
            "Success", 
            f"Processing Complete!\n\n"
            f"Total cells processed: {processed_cells}\n"
            f"Red cells (Absent + Overtime > 0): {red_cells}\n"
            f"Green cells (Absent + Overtime = 0): {green_cells}\n\n"
            f"Output file: {os.path.basename(output_path)}"
        )
        
        return output_path
        
    except FileNotFoundError as e:
        error_msg = f"File not found: {str(e)}"
        print(f"Error: {error_msg}")
        messagebox.showerror("File Error", error_msg)
        return None
        
    except pd.errors.EmptyDataError:
        error_msg = "The selected file appears to be empty or corrupted."
        print(f"Error: {error_msg}")
        messagebox.showerror("Data Error", error_msg)
        return None
        
    except KeyError as e:
        error_msg = f"Required column not found: {str(e)}\nPlease check your file format."
        print(f"Error: {error_msg}")
        messagebox.showerror("Column Error", error_msg)
        return None
        
    except Exception as e:
        error_msg = f"An unexpected error occurred: {str(e)}"
        print(f"Error: {error_msg}")
        messagebox.showerror("Error", error_msg)
        return None
        
    finally:
        root.destroy()  # Clean up tkinter

if __name__ == "__main__":
    try:
        output_file = main()
        if output_file:
            print(f"\nSuccess! Check the output file: {output_file}")
        else:
            print("\nProcess was cancelled or failed.")
    except Exception as e:
        print(f"\nFatal error occurred: {str(e)}")
    
    input("Press Enter to exit...")